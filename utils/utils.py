import pandas as pd
import re
from openpyxl import load_workbook
from PEMA.models import Articulo
from PEMA.models import Maestro
from PEMA.models import Materia
from PEMA.models import Prestatario


def importar_material(file_path):
    """
    Importa los datos de la lista de material.
    :param file_path:
    :return:
    """
    wb = load_workbook(file_path)
    sheet = wb.active

    # encontrar el indice de la ultima fila con datos en la columna de nombres (refactorizar esto)
    ultima_fila = sheet.max_row
    for fila in range(sheet.max_row, 1, -1):
        # verificar si la celda en la columna no esta vacia
        if sheet.cell(row=fila, column=3).value is not None:
            ultima_fila = fila
            break

    # leer SOLO las filas necesarias hasta la ultima fila con datos
    df = pd.read_excel(file_path, nrows=ultima_fila)

    for index, row in df.iterrows():
        # obtener los datos relevantes
        num_control = row['NÚMERO DE CONTROL']
        num_serie = row['NÚMERO DE SERIE']
        nombre_articulo = row['NOMBRE Y MODELO DEL EQUIPO']

        articulo, creado = Articulo.objects.get_or_create(
            nombre=nombre_articulo
        )

        # crear unidad
        unidad, creada = articulo.crear_unidad(num_control=num_control, num_serie=num_serie)

def importar_personal(file_path):
    """
    Importa la lista de personal del CEPA
    :param file_path: [ath del archivo xlsx
    :return:
    """
    df = pd.read_excel(file_path, header=1)

    for index, row in df.iterrows():
        username = row['NUMERO DE EMPLEADO']
        first_name = row['PERSONAL DEL CEPA']

    maestro, creado = Maestro.crear_usuario(
        username=username,
        first_name=first_name
    )

def importar_listas(file_path):
    # Cargar el archivo una vez
    df = pd.read_excel(file_path)

    """Parsear maestro"""
    # Obtener el num de empleado y nombre
    no_empleado = df.iloc[17, 1]
    nombre_empleado = df.iloc[17, 3]

    """Parsear materia"""
    # Obtener el nombre de la materia y la fecha
    nombre_materia = df.iloc[13, 3]
    fecha = df.iloc[5, 38]

    # dividir la fecha utilizando "/"
    partes_fecha = fecha.split("/")
    # obtener el anno (la tercera parte)
    anno = int(partes_fecha[2])
    if int(partes_fecha[1]) < 7:
        semestre = 1
    else:
        semestre = 2

    """Parsear prestatarios"""
    # Para evitar que pandas detecte 'unnamed n' columnas, releemos el
    # archivo para usar usecols en read_excel().
    df = pd.read_excel(file_path, header=21, usecols=['NOMBRE DEL ALUMNO', 'MATRÍCULA'])

    # Eliminar filas vacias o NAN
    df = df.dropna(subset=['NOMBRE DEL ALUMNO'], how='all')

    # Definir patron de expresion regular para encontrar numeros seguidos de dos espacios
    patron = r'\d+\s\s'

    # Encontrar el índice de la primera fila que contiene "Fecha/Hora"
    final = df[df['NOMBRE DEL ALUMNO'] == 'Fecha/Hora'].index[0]

    nombres = []
    matriculas = []

    # Iterar sobre los datos completos
    for index, row in df.loc[:final - 1].iterrows():
        nombre = row['NOMBRE DEL ALUMNO']
        matricula = row['MATRÍCULA']

        # Reemplazar el patron definido por una cadena vacia
        nombre = re.sub(patron, '', nombre)
        nombres.append(nombre)
        matriculas.append(int(matricula))

def parsear_listas(nombre_materia, anno, semestre, no_empleado, nombre_empleado, nombres, matriculas):
    """Crear Materia"""
    materia = Materia.objects.get_or_create(nombre=nombre_materia, year=anno, semestre=semestre)[0]

    """Crear maestro"""
    maestro = Maestro.crear_usuario(username=no_empleado, first_name=nombre_empleado, password=str(no_empleado))
    # agregar maestro a la materia
    materia.agregar_maestro(maestro)

    """Crear Prestatarios"""
    for i in nombres:
        prestatario = Prestatario.crear_usuario(username=matriculas[i], first_name=nombres[i], password=str(matriculas[i]))
        # agregar prestatario a materia
        materia.agregar_alumno(prestatario)


'''
# pathInventario = r"C:\Users\Hecta\Documents\test pandas\Inventario CEPA (Actualizado nov 2023).xlsx"
file = 'Lista de personal del CEPA (docentes quienes piden).xlsx'

dataPersonal = pd.read_excel(file)

for index, row in dataPersonal.iterrows():
    # obtener datos de la fila
    nombre = row['PERSONAL DEL CEPA']
    matricula = row['NUMERO DE EMPLEADO']

    # crear usuario prestatario
    user = Prestatario.crear_usuario(username=matricula, password=matricula)

    """
    # asignar usuario al grupo 'prestatarios'
    group, _ = Prestatario.crear_grupo()
    group.user_set.add(user)

    # otorgar permisos al usuario prestatario
    # esto es un ejemplo (revisar como se hace el mappeo):
    permisos = Permission.objects.filter(codename__in=['add_carrito', 'add_orden', 'add_corresponsableorden'])
    user.user_permissions.add(*permisos)
    """
'''

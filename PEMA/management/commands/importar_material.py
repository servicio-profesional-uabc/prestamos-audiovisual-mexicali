import pandas as pd
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from PEMA.models import Articulo
from PEMA.models import Unidad


class Command(BaseCommand):
    help = 'Importa datos desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Ruta del archivo Excel.')

    def handle(self, *args, **options):
        file_path = options['file_path']
        self.importar_material(file_path)

    def importar_material(self, file_path):
        """
        Importa los datos de la lista de material.
        :param file_path:
        :return:
        """
        wb = load_workbook(file_path)
        sheet = wb.active

        # encontrar el indice de la ultima fila con datos en la columna de nombres (refactorizar esto)
        ultima_fila = sheet.max_row
        for fila in range(sheet.max_row, 1, -1):
            # verificar si la celda en la columna no esta vacia
            if sheet.cell(row=fila, column=3).value is not None:
                ultima_fila = fila
                break

        # leer SOLO las filas necesarias hasta la ultima fila con datos
        df = pd.read_excel(file_path, nrows=ultima_fila)

        for index, row in df.iterrows():
            # obtener los datos relevantes
            num_control = row['NÚMERO DE CONTROL']
            num_serie = row['NÚMERO DE SERIE']
            nombre_articulo = row['NOMBRE Y MODELO DEL EQUIPO']
            """
            articulo, creado = Articulo.objects.get_or_create(nombre=nombre_articulo)

            # crear unidad
            unidad, creada = articulo.crear_unidad(num_control=num_control, num_serie=num_serie)
            """

            # buscar el articulo existente por nombre
            articulo, creado = Articulo.objects.get_or_create(nombre=nombre_articulo)

            # verificar si la unidad ya existe para el articulo y num_control
            try:
                unidad = Unidad.objects.get(articulo=articulo, num_control=num_control, num_serie=num_serie)
            except Unidad.DoesNotExist:
                # crear unidad solo si no existe
                if num_serie:
                    # buscar una unidad existente que tenga el num_serie
                    try:
                        unidad = Unidad.objects.get(articulo=articulo, num_serie=num_serie, num_control=num_control)
                    except Unidad.DoesNotExist:
                        # crear una nueva unidad solo si no existe con el mismo num_serie
                        unidad = Unidad.objects.create(articulo=articulo, num_control=num_control, num_serie=num_serie)
                else:
                    # si no hay num_serie, usar solo num_control para la unicidad
                    unidad, creada = articulo.crear_unidad(num_control=num_control, num_serie=num_serie)
